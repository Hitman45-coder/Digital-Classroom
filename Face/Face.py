import os
import sys
import cv2
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from insightface.app import FaceAnalysis
from collections import defaultdict
import threading

from PySide6.QtWidgets import (
    QApplication, QDialog, QFileDialog, QLabel, QDialogButtonBox, 
    QVBoxLayout, QInputDialog, QMessageBox, QTableWidgetItem, QHeaderView, QHBoxLayout, QSizePolicy,
    QWidget
)
from PySide6.QtCore import Qt, Signal, QThread, Slot, QRect
from PySide6.QtGui import QImage, QPixmap, QFont
from ui_main_window import Ui_Start_2
import shutil
import smtplib
from email.message import EmailMessage
import json
from dotenv import load_dotenv

load_dotenv()  # Loads the variable for Sever 

SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")

ip_url = "http://192.168.137.69:4747/video" 
ip_url1 = "http://192.168.137.253:4747/video"

# Manage path
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

model_path = resource_path("insightface/models")
os.environ["INSIGHTFACE_HOME"] = model_path

# Initialize & pass model path 
os.environ["INSIGHTFACE_DOWNLOAD_PROGRESS"] = "False"
face_recognition = FaceAnalysis(allowed_modules=['detection', 'recognition'], root=model_path)
face_recognition.prepare(ctx_id=-1, det_size=(640, 640))  # Set ctx_id to 0 for GPU

# Load existing(load) face encodings
known_face_encodings = defaultdict(list)
known_faces_dir = 'known_faces'
os.makedirs(known_faces_dir, exist_ok=True)
for person_name in os.listdir(known_faces_dir):
    person_dir = os.path.join(known_faces_dir, person_name)
    if not os.path.isdir(person_dir):
        continue
    for filename in os.listdir(person_dir):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            path = os.path.join(person_dir, filename)
            img = cv2.imread(path)
            faces = face_recognition.get(img)
            if faces:
                embedding = faces[0].embedding
                embedding = embedding / np.linalg.norm(embedding)
                known_face_encodings[person_name].append(embedding)

# Create .xlsx file to save attendance
excel_filename = 'attendance.xlsx'
try:
    wb = load_workbook(excel_filename)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date", "Time"])

class VideoThread(QThread):
    frame_data = Signal(int, QImage) # Emit id, Image
    attendance_updated = Signal(str, str, str)  # Emit name, date, and time

    def __init__(self, camera_id, url, parent=None):
        super().__init__(parent)
        self.running = False
        self.paused = False
        self.camera_id = camera_id
        self.url = url
        self.frame = None
        self.frame_lock = threading.Lock()

    # Recognize face & mark attendance
    def run(self):
        self.running = True
        cap = cv2.VideoCapture(self.url)
        if not cap.isOpened():
            print(f"Failed to open camera stream: {self.url}")
            return
        
        print(f"Successfully opened stream: {self.url}")
        already_marked = set()

        while self.running:
            if not self.paused:
                ret, frame = cap.read()
                if not ret:
                    print(f"Error reading frame from {self.url}, retrying...")
                    continue

                faces = face_recognition.get(frame)
                if faces:
                    for face in faces:
                        embedding = face.embedding
                        embedding = embedding / np.linalg.norm(embedding)
                        best_match = "Unknown"
                        best_distance = float('inf')

                        for name, embeddings in known_face_encodings.items():
                            distances = np.linalg.norm(np.array(embeddings) - embedding, axis=1)
                            min_dist = np.min(distances)
                            if min_dist < best_distance:
                                best_distance = min_dist
                                best_match = name

                        threshold = 1.0
                        name = best_match if best_distance < threshold else "Unknown"
                        x1, y1, x2, y2 = map(int, face.bbox)
                        cv2.rectangle(frame, (x1, y1), (x2, y2), (125, 100, 0), 2)
                        cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)
                        if name != "Unknown" and name not in already_marked:
                            now = datetime.now()
                            date_str = now.strftime("%Y-%m-%d")
                            time_str = now.strftime("%H:%M:%S")
                            # Check for duplicates in the Excel file for today
                            has_duplicate = any(row[0] == name and row[1] == date_str for row in ws.iter_rows(min_row=2, values_only=True))
                            if not has_duplicate:
                                ws.append([name, date_str, time_str])
                                wb.save(excel_filename) 
                                already_marked.add(name)
                                self.attendance_updated.emit(name, date_str, time_str)

                with self.frame_lock:
                    self.frame = frame
                rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = rgb_frame.shape
                bytes_per_line = ch * w
                qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                self.frame_data.emit(self.camera_id, qt_image)

        cap.release()
        wb.save(excel_filename)  # Save when thread stops
        print(f"Stopped stream: {self.url}")

    def stop(self):
        self.running = False
        self.wait()

    def pause(self):
        self.paused = True

    def resume(self):
        self.paused = False

    def get_latest_frame(self):
        with self.frame_lock:
            return self.frame.copy() if self.frame is not None else None

class MainApp(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(
            Qt.Window | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint
        )
        self.ui = Ui_Start_2()
        self.ui.setupUi(self)

        # Verify UI elements
        if not hasattr(self.ui, 'Start_Button') or not hasattr(self.ui, 'Stop_Button') or \
           not hasattr(self.ui, 'Mail_Button') or not hasattr(self.ui, 'Add_Button') or \
           not hasattr(self.ui, 'Remove_Button') or not hasattr(self.ui, 'Attendance') or \
           not hasattr(self.ui, 'video_container'):
            raise AttributeError("One or more UI elements are missing in Ui_Start_2. Check ui_main_window.py or the .ui file.")

        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(main_layout)

        self.video_widget = self.ui.video_container  
        self.video_layout = QVBoxLayout(self.video_widget)
        self.video_layout.setContentsMargins(0, 0, 0, 0)
        self.video_layout.setSpacing(0)

        button_panel = QWidget()
        button_layout = QVBoxLayout(button_panel)
        button_layout.addWidget(self.ui.Start_Button)
        button_layout.addWidget(self.ui.Stop_Button)
        button_layout.addWidget(self.ui.Mail_Button)
        button_layout.addWidget(self.ui.Add_Button)
        button_layout.addWidget(self.ui.Remove_Button)
        button_layout.addWidget(self.ui.Attendance, stretch=1)  
        button_layout.addStretch()
        button_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(self.video_widget)
        main_layout.addWidget(button_panel)

        self.present_today = {}

        self.ui.Attendance.setColumnCount(3)
        self.ui.Attendance.setHorizontalHeaderLabels(["Name", "Date", "Time"])
        self.ui.Attendance.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ui.Attendance.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.ui.Attendance.setMinimumHeight(200)  

        self.camera_threads = {}
        self.camera_views = {}

        # Explicit signal-slot connections
        self.ui.Start_Button.clicked.connect(self.start_attendance)
        self.ui.Stop_Button.clicked.connect(self.stop_attendance)
        self.ui.Add_Button.clicked.connect(self.add_face)
        self.ui.Remove_Button.clicked.connect(self.remove_face)
        self.ui.Mail_Button.clicked.connect(self.send_absentee_emails)

        self.update_today_present_list()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        for cam_id, label in self.camera_views.items():
            self._adjust_label_size(label)

    def _adjust_label_size(self, label):
        if not self.video_widget or not self.video_layout.count():
            return
        available_width = self.video_widget.width()
        available_height = self.video_widget.height() // len(self.camera_views)
        label.setMaximumSize(available_width, available_height)
        label.setMinimumSize(available_width, available_height)

    @Slot(int, QImage)
    def update_frame(self, cam_id, image):
        pixmap = QPixmap.fromImage(image)
        if cam_id in self.camera_views:
            label = self.camera_views[cam_id]
            scaled_pixmap = pixmap.scaled(label.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation)
            label.setPixmap(scaled_pixmap)

    @Slot(str, str, str)
    def on_attendance_updated(self, name, date_str, time_str):
        if name not in self.present_today:
            self.present_today[name] = (date_str, time_str)
            self.update_today_present_list()
            wb.save(excel_filename)  # Save workbook after update

    def start_attendance(self):
        if self.camera_threads:
            return

        for i in reversed(range(self.video_layout.count())):
            widget = self.video_layout.itemAt(i).widget()
            if widget:
                self.video_layout.removeWidget(widget)
                widget.deleteLater()

        self.camera_views.clear()
        self.camera_threads.clear()

        urls = [(0, ip_url), (1, ip_url1)]
        for cam_id, url in urls:
            thread = VideoThread(cam_id, url, self)
            thread.frame_data.connect(self.update_frame)
            thread.attendance_updated.connect(self.on_attendance_updated)
            self.camera_threads[cam_id] = thread

            label = QLabel()
            label.setStyleSheet("border: 1px solid gray;")
            label.setAlignment(Qt.AlignCenter)
            label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.video_layout.addWidget(label)
            self.camera_views[cam_id] = label
            self._adjust_label_size(label)
            thread.start()

        self.update_today_present_list()

    def stop_attendance(self):
        for thread in self.camera_threads.values():
            thread.stop()
        self.camera_threads.clear()

        for label in self.camera_views.values():
            self.video_layout.removeWidget(label)
            label.deleteLater()
        self.camera_views.clear()

        self.update_today_present_list()

    def update_today_present_list(self):
        today = datetime.now().strftime("%Y-%m-%d")
        self.present_today.clear()

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, date_str, time_str = row
            if date_str == today:
                self.present_today[name] = (date_str, time_str)

        self.ui.Attendance.setRowCount(len(self.present_today))
        row = 0
        for name, (date_str, time_str) in sorted(self.present_today.items()):
            item_name = QTableWidgetItem(name)
            item_name.setFlags(item_name.flags() & ~Qt.ItemIsEditable)
            self.ui.Attendance.setItem(row, 0, item_name)

            item_date = QTableWidgetItem(date_str)
            item_date.setFlags(item_date.flags() & ~Qt.ItemIsEditable)
            self.ui.Attendance.setItem(row, 1, item_date)

            item_time = QTableWidgetItem(time_str)
            item_time.setFlags(item_time.flags() & ~Qt.ItemIsEditable)
            self.ui.Attendance.setItem(row, 2, item_time)
            row += 1

    def add_face(self):
        method, ok = QInputDialog.getItem(
            self, "Select Method", "Add face by:",
            ["Capture from Phone Camera", "Upload from Disk"],
            0, False
        )
        if not ok:
            return

        name, ok = QInputDialog.getText(self, "Person's Name", "Enter person's name:")
        if not ok or not name.strip():
            return

        name = name.strip()

        email, ok = QInputDialog.getText(self, "Email Address", f"Enter email for {name}:")
        if not ok or not email.strip():
            return
        email = email.strip()

        self.save_email(name, email)

        folder_path = os.path.join("known_faces", name)
        os.makedirs(folder_path, exist_ok=True)

        if method == "Capture from Phone Camera":
            if not self.camera_threads:
                QMessageBox.warning(self, "Error", "No camera feeds are running. Start attendance first.")
                return

            cam_options = {0: "Camera 1 (Realme)", 1: "Camera 2 (Mi)"}
            cam_id, ok = QInputDialog.getItem(self, "Select Camera", "Choose camera to capture face:",
                                            list(cam_options.values()), 0, False)
            if not ok:
                return
            selected_cam_id = [k for k, v in cam_options.items() if v == cam_id][0]

            other_cam_id = 1 - selected_cam_id
            if other_cam_id in self.camera_threads:
                self.camera_threads[other_cam_id].pause()

            capture_dialog = QDialog(self)
            capture_dialog.setWindowTitle(f"Capture Face from {cam_options[selected_cam_id]}")
            layout = QVBoxLayout(capture_dialog)
            capture_label = QLabel()
            capture_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(capture_label)

            buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
            layout.addWidget(buttons)
            buttons.accepted.connect(capture_dialog.accept)
            buttons.rejected.connect(capture_dialog.reject)

            def update_capture_frame():
                frame = self.camera_threads[selected_cam_id].get_latest_frame()
                if frame is not None:
                    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    h, w, ch = rgb_frame.shape
                    bytes_per_line = ch * w
                    qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                    pixmap = QPixmap.fromImage(qt_image).scaled(640, 480, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    capture_label.setPixmap(pixmap)

            timer = self.startTimer(100)
            def timer_event():
                update_capture_frame()

            capture_dialog.timer_event = timer_event
            capture_dialog.timer_id = timer

            if capture_dialog.exec() == QDialog.Accepted:
                frame = self.camera_threads[selected_cam_id].get_latest_frame()
                if frame is not None:
                    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    height, width, _ = frame_rgb.shape
                    bytes_per_line = 3 * width
                    preview_img = QImage(frame_rgb.data, width, height, bytes_per_line, QImage.Format_RGB888)
                    preview_dialog = QDialog(self)
                    preview_dialog.setWindowTitle("Confirm Photo")
                    preview_layout = QVBoxLayout(preview_dialog)
                    preview_label = QLabel()
                    preview_label.setPixmap(QPixmap.fromImage(preview_img).scaled(400, 400, Qt.KeepAspectRatio))
                    preview_layout.addWidget(preview_label)
                    preview_buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                    preview_layout.addWidget(preview_buttons)
                    preview_buttons.accepted.connect(preview_dialog.accept)
                    preview_buttons.rejected.connect(preview_dialog.reject)

                    if preview_dialog.exec() == QDialog.Accepted:
                        count = len([f for f in os.listdir(folder_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
                        save_path = os.path.join(folder_path, f"{count + 1}.jpg")
                        cv2.imwrite(save_path, frame)
                        QMessageBox.information(self, "Success", f"Image saved to {save_path}")
                        self.refresh_known_encodings()
                    else:
                        QMessageBox.information(self, "Cancelled", "Face not saved.")
                else:
                    QMessageBox.warning(self, "Error", "Failed to capture image.")
            else:
                QMessageBox.information(self, "Cancelled", "Face capture cancelled.")

            if other_cam_id in self.camera_threads:
                self.camera_threads[other_cam_id].resume()
            self.killTimer(capture_dialog.timer_id)
        else:
            file_path, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Images (*.png *.jpg *.jpeg)")
            if not file_path:
                return
            frame = cv2.imread(file_path)

            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            height, width, _ = frame_rgb.shape
            bytes_per_line = 3 * width
            preview_img = QImage(frame_rgb.data, width, height, bytes_per_line, QImage.Format_RGB888)

            dialog = QDialog(self)
            dialog.setWindowTitle("Confirm Photo")
            layout = QVBoxLayout(dialog)
            label = QLabel()
            label.setPixmap(QPixmap.fromImage(preview_img).scaled(400, 400, Qt.KeepAspectRatio))
            layout.addWidget(label)
            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            layout.addWidget(buttons)
            buttons.accepted.connect(dialog.accept)
            buttons.rejected.connect(dialog.reject)

            if dialog.exec() == QDialog.Accepted:
                count = len([f for f in os.listdir(folder_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
                save_path = os.path.join(folder_path, f"{count + 1}.jpg")
                cv2.imwrite(save_path, frame)
                QMessageBox.information(self, "Success", f"Image saved to {save_path}")
                self.refresh_known_encodings()
            else:
                QMessageBox.information(self, "Cancelled", "Face not saved.")
            self.refresh_known_encodings()

    def timerEvent(self, event):
        if hasattr(self.sender(), 'timer_event'):
            self.sender().timer_event()

    def remove_face(self):
        name, ok = QInputDialog.getText(self, "Remove Face", "Enter person's name to delete:")
        if not ok or not name.strip():
            return
        name = name.strip()
        folder_path = os.path.join("known_faces", name)
        if not os.path.exists(folder_path):
            QMessageBox.warning(self, "Error", "Person not found.")
            return
        confirm = QMessageBox.question(self, "Confirm Deletion", f"Delete all data for {name}?", QMessageBox.Yes | QMessageBox.No)
        if confirm == QMessageBox.Yes:
            shutil.rmtree(folder_path)
            self.delete_email(name)
            QMessageBox.information(self, "Deleted", f"All data for {name} deleted.")
            self.refresh_known_encodings()

    def refresh_known_encodings(self):
        global known_face_encodings
        known_face_encodings = defaultdict(list)
        for person_name in os.listdir(known_faces_dir):
            person_dir = os.path.join(known_faces_dir, person_name)
            if not os.path.isdir(person_dir):
                continue
            for filename in os.listdir(person_dir):
                if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                    path = os.path.join(person_dir, filename)
                    img = cv2.imread(path)
                    faces = face_recognition.get(img)
                    if faces:
                        embedding = faces[0].embedding
                        embedding = embedding / np.linalg.norm(embedding)
                        known_face_encodings[person_name].append(embedding)

    def get_absentees_today(self):
        all_people = set(os.listdir("known_faces"))
        present_today = set(self.present_today.keys())  
        return list(all_people - present_today)

    def save_email(self, name, email):
        email_file = "emails.json"
        if os.path.exists(email_file):
            with open(email_file, "r") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    data = {}
        else:
            data = {}
        data[name] = email
        with open(email_file, "w") as f:
            json.dump(data, f, indent=4)

    def delete_email(self, name):
        email_file = "emails.json"
        if os.path.exists(email_file):
            with open(email_file, "r") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    data = {}
            if name in data:
                del data[name]
                with open(email_file, "w") as f:
                    json.dump(data, f, indent=4)

    def send_absentee_emails(self):
        absentees = self.get_absentees_today()
        if not absentees:
            QMessageBox.information(self, "Info", "No absentees today!")
            return
        email_file = "emails.json"
        if os.path.exists(email_file):
            with open(email_file, "r") as f:
                try:
                    emails = json.load(f)
                except json.JSONDecodeError:
                    emails = {}
        else:
            emails = {}
        SMTP_SERVER = 'smtp.gmail.com'
        SMTP_PORT = 587
        
        try:
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
        except Exception as e:
            QMessageBox.warning(self, "SMTP Error", f"Failed to connect to SMTP server:\n{e}")
            return
        for name in absentees:
            recipient_email = emails.get(name)
            if not recipient_email:
                print(f"No email found for {name}, skipping.")
                continue
            msg = EmailMessage()
            msg['Subject'] = "Attendance Alert"
            msg['From'] = SMTP_USER
            msg['To'] = recipient_email
            msg.set_content(f"Dear {name},\n\nOur records show you were absent today ({datetime.now().strftime('%Y-%m-%d')}). Please contact your supervisor if this is an error.\n\nBest regards,\nAttendance System")
            try:
                server.send_message(msg)
                print(f"Sent absentee email to {name} at {recipient_email}")
            except Exception as e:
                print(f"Failed to send email to {name} ({recipient_email}): {e}")
        server.quit()
        QMessageBox.information(self, "Emails Sent", "Absentee emails have been sent.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec())
